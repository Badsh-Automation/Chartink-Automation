#!/usr/bin/env python3
"""
================================================================================
CHARTINK DASHBOARD - OUTLOOK/EMAIL VERSION
================================================================================
Works with: Gmail, Outlook, Office 365, Yahoo, etc.
================================================================================
"""

import os
import sys
import time
import re
import smtplib
from datetime import date, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

# CONFIG
CHARTINK_EMAIL = os.environ.get("CHARTINK_EMAIL", "")
CHARTINK_PASSWORD = os.environ.get("CHARTINK_PASSWORD", "")
BASE_DIR = os.path.join(os.getcwd(), "data")
CHARTINK1_FOLDER = os.path.join(BASE_DIR, "Chartink1")
CHARTINK_FOLDER = os.path.join(BASE_DIR, "Chartink")
OUTPUT_FOLDER = os.path.join(os.getcwd(), "output")

# Email config from secrets
EMAIL_SENDER = os.environ.get("EMAIL_SENDER", "")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD", "")
EMAIL_RECEIVER = os.environ.get("EMAIL_RECEIVER", "")
EMAIL_PROVIDER = os.environ.get("EMAIL_PROVIDER", "gmail").lower()  # gmail, outlook, yahoo

SCREENER_LONG_URL = "https://chartink.com/screener/badsha-long"
SCREENER_SHORT_URL = "https://chartink.com/screener/badsha-short"
LONG_PREFIX = "badsha long"
SHORT_PREFIX = "badsha short"

def is_weekday():
    return datetime.now().weekday() < 5

# Weekend skip toggle - Set SKIP_WEEKEND=false in GitHub Secrets to run on Sat/Sun
SKIP_WEEKEND = os.environ.get("SKIP_WEEKEND", "true").strip().lower() in ("true", "1", "yes", "on")

# SMTP Config based on provider
SMTP_CONFIG = {
    "gmail": {"server": "smtp.gmail.com", "port": 587},
    "outlook": {"server": "smtp.office365.com", "port": 587},
    "office365": {"server": "smtp.office365.com", "port": 587},
    "yahoo": {"server": "smtp.mail.yahoo.com", "port": 587},
    "hotmail": {"server": "smtp.office365.com", "port": 587},
}

FNO_STOCKS = [
    "IBULHSGFIN", "MCDOWELL-N", "L&TFH", "GMRINFRA", "TATAMOTORS", "IDFC",
    "LTIM", "PEL", "IDEA", "ABCAPITAL", "IDFCFIRSTB", "HDFCBANK", "IRCTC",
    "SAIL", "TATASTEEL", "IOC", "BANDHANBNK", "CANBK", "INFY", "AUBANK",
    "BEL", "MOTHERSON", "POWERGRID", "ONGC", "KOTAKBANK", "PNB", "GAIL",
    "NMDC", "ICICIBANK", "BANKBARODA", "ASHOKLEY", "ITC", "BHARTIARTL",
    "WIPRO", "VEDL", "RELIANCE", "NTPC", "SBIN", "HINDCOPPER", "NATIONALUM",
    "BPCL", "JUBLFOOD", "HCLTECH", "ZEEL", "PETRONET", "BHEL", "COALINDIA",
    "BAJFINANCE", "RECLTD", "INDUSTOWER", "PFC", "FEDERALBNK", "MANAPPURAM",
    "AXISBANK", "SHRIRAMFIN", "HINDALCO", "HINDPETRO", "TCS", "IEX", "INDHOTEL",
    "RBLBANK", "SBICARD", "BIOCON", "UPL", "DLF", "M&M", "INDUSINDBK",
    "DRREDDY", "HDFCLIFE", "COFORGE", "TATAPOWER", "HINDUNILVR", "SUNPHARMA",
    "BSOFT", "CIPLA", "TECHM", "MCX", "ASIANPAINT", "MARICO", "NAUKRI",
    "AMBUJACEM", "DABUR", "LT", "ICICIPRULI", "HAL", "JSWSTEEL", "CROMPTON",
    "TATACONSUM", "CUB", "CHOLAFIN", "AUROPHARMA", "CONCOR", "LAURUSLABS",
    "VOLTAS", "ABFRL", "NESTLEIND", "JINDALSTEL", "PIDILITIND", "ADANIPORTS",
    "LICHSGFIN", "HAVELLS", "SUNTV", "GODREJCP", "ADANIENT", "TVSMOTOR",
    "BHARATFORG", "SBILIFE", "MUTHOOTFIN", "DIXON", "BAJAJFINSV", "GODREJPROP",
    "DELTACORP", "INDIGO", "HDFCAMC", "GRASIM", "IGL", "TITAN", "PERSISTENT",
    "ZYDUSLIFE", "BERGEPAINT", "APOLLOTYRE", "TRENT", "TATACOMM", "M&MFIN",
    "LUPIN", "CUMMINSIND", "POLYCAB", "EXIDEIND", "COROMANDEL", "SYNGENE",
    "MPHASIS", "SIEMENS", "ICICIGI", "OBEROIRLTY", "DALBHARAT", "EICHERMOT",
    "ASTRAL", "APOLLOHOSP", "SRF", "GLENMARK", "ABB", "LALPATHLAB", "HEROMOTOCO",
    "COLPAL", "CHAMBLFERT", "MFSL", "INDIACEM", "AARTIIND", "GRANULES", "MARUTI",
    "TATACHEM", "BALRAMCHIN", "DIVISLAB", "MGL", "TORNTPHARM", "PIIND",
    "BRITANNIA", "LTTS", "CANFINHOME", "ULTRACEMCO", "NAVINFLUOR", "GUJGASLTD",
    "OFSS", "BAJAJ-AUTO", "IPCALAB", "PVRINOX", "ACC", "BALKRISIND", "RAMCOCEM",
    "ALKEM", "DEEPAKNTR", "BATAINDIA", "GNFC", "PAGEIND", "JKCEMENT", "ESCORTS",
    "INDIAMART", "SHREECEM", "ATUL", "UBL", "BOSCHLTD", "METROPOLIS", "MRF",
    "ABBOTINDIA"
]

def create_all_folders():
    for folder in [CHARTINK1_FOLDER, CHARTINK_FOLDER, OUTPUT_FOLDER]:
        os.makedirs(folder, exist_ok=True)

def download_files():
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    def get_driver():
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        prefs = {
            "download.default_directory": os.path.expanduser("~"),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)
        return webdriver.Chrome(options=options)

    def login(driver):
        driver.get("https://chartink.com/login")
        wait = WebDriverWait(driver, 20)
        time.sleep(3)
        for by, sel in [(By.ID, "login-email"), (By.NAME, "email"), (By.CSS_SELECTOR, "input[type='email']")]:
            try:
                el = wait.until(EC.presence_of_element_located((by, sel)))
                el.clear(); el.send_keys(CHARTINK_EMAIL); break
            except: continue
        for by, sel in [(By.ID, "login-password"), (By.NAME, "password"), (By.CSS_SELECTOR, "input[type='password']")]:
            try:
                el = driver.find_element(by, sel)
                el.clear(); el.send_keys(CHARTINK_PASSWORD); break
            except: continue
        for by, sel in [(By.XPATH, "//button[.//span[contains(text(),'Log in')]]"), (By.CSS_SELECTOR, "button.primary-button"), (By.CSS_SELECTOR, "button[type='submit']")]:
            try:
                btn = driver.find_element(by, sel)
                driver.execute_script("arguments[0].click();", btn); break
            except: continue
        time.sleep(5)
        print("[LOGIN] Done")

    def set_download_dir(driver, path):
        os.makedirs(path, exist_ok=True)
        driver.execute_cdp_cmd("Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": path})

    def wait_for_file(directory, before, timeout=30):
        end = time.time() + timeout
        while time.time() < end:
            curr = set(os.listdir(directory))
            cand = [f for f in (curr - before) if not f.endswith((".crdownload", ".tmp"))]
            if cand:
                fp = max([os.path.join(directory, f) for f in cand], key=os.path.getmtime)
                s1 = os.path.getsize(fp); time.sleep(1)
                if os.path.getsize(fp) == s1: return fp
            time.sleep(1)
        return None

    def rename_file(filepath, prefix):
        if not filepath or not os.path.exists(filepath): return None
        directory = os.path.dirname(filepath)
        _, ext = os.path.splitext(filepath)
        date_str = date.today().strftime("%Y-%m-%d")
        new_name = f"{prefix} ({date_str}){ext}"
        new_path = os.path.join(directory, new_name)
        c = 1
        while os.path.exists(new_path):
            new_name = f"{prefix} ({date_str})_{c}{ext}"
            new_path = os.path.join(directory, new_name); c += 1
        os.rename(filepath, new_path)
        return new_path

    def download_one(driver, url, folder, prefix):
        set_download_dir(driver, folder)
        before = set(os.listdir(folder))
        driver.get(url); wait = WebDriverWait(driver, 25); time.sleep(5)
        btn = None
        for by, sel in [
            (By.XPATH, "//button[@aria-label='Excel']"),
            (By.CSS_SELECTOR, "button[aria-label='Excel']"),
            (By.XPATH, "//i[contains(@class,'fa-file-excel')]/ancestor::button"),
            (By.XPATH, "//button[contains(text(),'Export')]"),
            (By.CSS_SELECTOR, ".buttons-csv"),
            (By.XPATH, "//*[contains(@class,'fa-download')]"),
        ]:
            try:
                btn = wait.until(EC.element_to_be_clickable((by, sel)))
                break
            except: continue
        if btn:
            driver.execute_script("arguments[0].click();", btn)
            print(f"[DOWNLOAD] {prefix} triggered")
            nf = wait_for_file(folder, before, 30)
            if nf:
                rf = rename_file(nf, prefix)
                print(f"[SUCCESS] {rf}"); return True
        print(f"[WARNING] {prefix} download issue")
        return False

    driver = get_driver()
    try:
        login(driver)
        download_one(driver, SCREENER_LONG_URL, CHARTINK1_FOLDER, LONG_PREFIX)
        download_one(driver, SCREENER_SHORT_URL, CHARTINK_FOLDER, SHORT_PREFIX)
    finally:
        driver.quit()

def transform_file(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='Watchlist', engine='openpyxl')
        if 'Add Column' in df.columns:
            df = df.drop(columns=['Add Column'])
        return df
    except Exception as e:
        print(f"[ERROR] {os.path.basename(file_path)}: {e}")
        return pd.DataFrame()

def process_chartink1_folder():
    print(f"\n[INFO] LONG scan: {CHARTINK1_FOLDER}")
    if not os.path.exists(CHARTINK1_FOLDER): return pd.DataFrame()
    files = [f for f in os.listdir(CHARTINK1_FOLDER) if f.endswith('.xlsx') and not f.startswith('~$')]
    if not files: return pd.DataFrame()
    print(f"[INFO] Found {len(files)} files: {files}")
    dfs = []
    for f in files:
        df = transform_file(os.path.join(CHARTINK1_FOLDER, f))
        if not df.empty:
            df['Source.Name'] = f
            dfs.append(df)
    if not dfs: return pd.DataFrame()
    combined = pd.concat(dfs, ignore_index=True)
    if '%_change' in combined.columns:
        combined = combined[combined['%_change'] >= 3].copy()
        combined['Signal'] = combined['%_change'].apply(lambda x: 'BUY' if pd.to_numeric(x, errors='coerce') >= 5 else 'SELL')
    print(f"[INFO] Long: {len(combined)} rows")
    return combined

def process_chartink_folder():
    print(f"\n[INFO] SHORT scan: {CHARTINK_FOLDER}")
    if not os.path.exists(CHARTINK_FOLDER): return pd.DataFrame()
    files = [f for f in os.listdir(CHARTINK_FOLDER) if f.endswith('.xlsx') and not f.startswith('~$')]
    if not files: return pd.DataFrame()
    print(f"[INFO] Found {len(files)} files: {files}")
    dfs = []
    for f in files:
        df = transform_file(os.path.join(CHARTINK_FOLDER, f))
        if not df.empty:
            df['Source.Name'] = f
            dfs.append(df)
    if not dfs: return pd.DataFrame()
    combined = pd.concat(dfs, ignore_index=True)
    if '%_change' in combined.columns:
        combined = combined[combined['%_change'] >= 1].copy()
        combined['Signal'] = combined['%_change'].apply(lambda x: 'BUY' if pd.to_numeric(x, errors='coerce') >= 2 else 'SELL')
    print(f"[INFO] Short: {len(combined)} rows")
    return combined

def create_long_sheet(chartink1_df):
    if chartink1_df.empty: return pd.DataFrame()
    df = chartink1_df.copy()
    df['Date'] = df['Source.Name'].apply(lambda x: re.search(r'\((\d{4}-\d{2}-\d{2})\)', str(x)).group(1) if re.search(r'\((\d{4}-\d{2}-\d{2})\)', str(x)) else '')
    return pd.DataFrame({
        'File Name': df['Source.Name'], 'Date': df['Date'], 'Stock Name': df['Stock Name'],
        'Symbol': df['Symbol'], 'Date ': df['Date'], 'Volume': df['volume'],
        'Changes': df['%_change'].apply(lambda x: f"{x}%"), 'Last Price': df['close'], 'Status': df['Signal']
    })

def create_short_sheet(chartink_df):
    if chartink_df.empty: return pd.DataFrame()
    df = chartink_df.copy()
    df['Date'] = df['Source.Name'].apply(lambda x: re.search(r'\((\d{4}-\d{2}-\d{2})\)', str(x)).group(1) if re.search(r'\((\d{4}-\d{2}-\d{2})\)', str(x)) else '')
    return pd.DataFrame({
        'File Name': df['Source.Name'], 'Date': df['Date'], 'Stock Name': df['Stock Name'],
        'Symbol': df['Symbol'], 'Volume': df['volume'],
        'Changes': df['%_change'].apply(lambda x: f"{x}%"), 'Last Price': df['close'], 'Status': df['Signal']
    })

def create_buy_sheet(long_df):
    if long_df.empty: return pd.DataFrame()
    buy_df = long_df[long_df['Status'] == 'BUY'].copy()
    buy_df = buy_df.reset_index(drop=True)
    print(f"[INFO] BUY sheet: {len(buy_df)} rows (only BUY)")
    return buy_df

def create_sell_sheet(short_df):
    if short_df.empty: return pd.DataFrame()
    sell_df = short_df[short_df['Status'] == 'SELL'].copy()
    sell_df = sell_df.reset_index(drop=True)
    print(f"[INFO] SELL sheet: {len(sell_df)} rows (only SELL)")
    return sell_df

def create_fno_sheet(chartink1_df, chartink_df):
    all_data = pd.concat([chartink1_df, chartink_df], ignore_index=True)
    if all_data.empty: return pd.DataFrame()
    fno_df = pd.DataFrame({'Symbol': all_data['Symbol'].unique()})
    fno_df['In FNO List'] = fno_df['Symbol'].apply(lambda x: 'YES' if x in FNO_STOCKS else 'NO')
    fno_df['FNO Category'] = fno_df['Symbol'].apply(lambda x: 'FNO Stock' if x in FNO_STOCKS else 'Non-FNO')
    symbol_counts = all_data['Symbol'].value_counts().to_dict()
    fno_df['Count in Data'] = fno_df['Symbol'].map(symbol_counts)
    return fno_df.sort_values(['FNO Category', 'Symbol'])

def append_to_excel(existing_file, new_data_dict):
    if not os.path.exists(existing_file):
        return False
    try:
        existing_sheets = {}
        with pd.ExcelFile(existing_file) as xls:
            for sheet_name in xls.sheet_names:
                existing_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
        with pd.ExcelWriter(existing_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for sheet_name, new_df in new_data_dict.items():
                if sheet_name in existing_sheets and not existing_sheets[sheet_name].empty:
                    existing_df = existing_sheets[sheet_name]
                    combined = pd.concat([existing_df, new_df], ignore_index=True)
                    combined = combined.drop_duplicates()
                    combined.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  ✓ {sheet_name}: Appended (total {len(combined)} rows)")
                else:
                    new_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  ✓ {sheet_name}: Created ({len(new_df)} rows)")
        return True
    except Exception as e:
        print(f"[WARNING] Could not append: {e}, creating fresh file")
        return False

def generate_dashboard(chartink1_df, chartink_df, long_df, short_df, buy_df, sell_df, fno_df):
    output_file = os.path.join(OUTPUT_FOLDER, 'Chartink Dashboard.xlsx')
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    print(f"\n[INFO] Generating: {output_file}")

    new_data = {
        'Chartink1 (2)': chartink1_df if not chartink1_df.empty else pd.DataFrame(),
        'Chartink (2)': chartink_df if not chartink_df.empty else pd.DataFrame(),
        'Long': long_df if not long_df.empty else pd.DataFrame(),
        'Short': short_df if not short_df.empty else pd.DataFrame(),
        'Stocks FNO': pd.DataFrame({'FNO STOCK LIST': FNO_STOCKS}),
        'BUY': buy_df if not buy_df.empty else pd.DataFrame(),
        'SELL': sell_df if not sell_df.empty else pd.DataFrame(),
        'FNO': fno_df if not fno_df.empty else pd.DataFrame(),
    }

    if os.path.exists(output_file):
        print("[INFO] Existing file found, appending data...")
        success = append_to_excel(output_file, new_data)
        if success:
            print("[SUCCESS] Data appended to existing file")
            return output_file

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        if not chartink1_df.empty:
            cols = ['Source.Name'] + [c for c in chartink1_df.columns if c != 'Source.Name']
            chartink1_df[cols].to_excel(writer, sheet_name='Chartink1 (2)', index=False)
            _format_sheet(writer.sheets['Chartink1 (2)'])
            print(f"  ✓ Chartink1 (2): {len(chartink1_df)} rows")

        if not chartink_df.empty:
            cols = ['Source.Name'] + [c for c in chartink_df.columns if c != 'Source.Name']
            chartink_df[cols].to_excel(writer, sheet_name='Chartink (2)', index=False)
            _format_sheet(writer.sheets['Chartink (2)'])
            print(f"  ✓ Chartink (2): {len(chartink_df)} rows")

        if not long_df.empty:
            long_df.to_excel(writer, sheet_name='Long', index=False)
            _format_sheet(writer.sheets['Long'])
            print(f"  ✓ Long: {len(long_df)} rows")

        if not short_df.empty:
            short_df.to_excel(writer, sheet_name='Short', index=False)
            _format_sheet(writer.sheets['Short'])
            print(f"  ✓ Short: {len(short_df)} rows")

        fno_list_df = pd.DataFrame({'FNO STOCK LIST': FNO_STOCKS})
        fno_list_df.to_excel(writer, sheet_name='Stocks FNO', index=False)
        _format_sheet(writer.sheets['Stocks FNO'])
        print(f"  ✓ Stocks FNO: {len(FNO_STOCKS)} stocks")

        if not buy_df.empty:
            buy_df.to_excel(writer, sheet_name='BUY', index=False)
            _format_sheet(writer.sheets['BUY'])
            _apply_buy_color(writer.sheets['BUY'])
            print(f"  ✓ BUY: {len(buy_df)} rows (only BUY)")

        if not sell_df.empty:
            sell_df.to_excel(writer, sheet_name='SELL', index=False)
            _format_sheet(writer.sheets['SELL'])
            _apply_sell_color(writer.sheets['SELL'])
            print(f"  ✓ SELL: {len(sell_df)} rows (only SELL)")

        if not fno_df.empty:
            fno_df.to_excel(writer, sheet_name='FNO', index=False)
            _format_sheet(writer.sheets['FNO'])
            print(f"  ✓ FNO: {len(fno_df)} rows")

    print(f"[SUCCESS] Dashboard: {output_file}")
    return output_file

def _format_sheet(worksheet):
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 40))
            except:
                pass
        worksheet.column_dimensions[column].width = max_length + 3
    worksheet.freeze_panes = 'A2'
    for row_idx in range(2, worksheet.max_row + 1):
        if row_idx % 2 == 0:
            for cell in worksheet[row_idx]:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def _apply_buy_color(worksheet):
    for row in range(2, worksheet.max_row + 1):
        for cell in worksheet[row]:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(color="006100")

def _apply_sell_color(worksheet):
    for row in range(2, worksheet.max_row + 1):
        for cell in worksheet[row]:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(color="9C0006")

# EMAIL FUNCTION - Works with Gmail, Outlook, Yahoo, etc.
def send_email(file_path):
    try:
        if not all([EMAIL_SENDER, EMAIL_PASSWORD, EMAIL_RECEIVER]):
            print("[WARNING] Email credentials not set, skipping email")
            return False

        # Get SMTP config for the provider
        smtp_config = SMTP_CONFIG.get(EMAIL_PROVIDER, SMTP_CONFIG["gmail"])
        smtp_server = smtp_config["server"]
        smtp_port = smtp_config["port"]

        print(f"[EMAIL] Using {EMAIL_PROVIDER} SMTP: {smtp_server}:{smtp_port}")

        msg = MIMEMultipart()
        msg['From'] = EMAIL_SENDER
        msg['To'] = EMAIL_RECEIVER
        msg['Subject'] = f'Chartink Dashboard - {datetime.now().strftime("%Y-%m-%d")}'

        body = f"""
Hello,

Your Chartink Dashboard for {datetime.now().strftime("%d-%m-%Y")} is ready!

Attached file contains:
• Chartink1 (2) - Long scan data
• Chartink (2) - Short scan data
• Long - Reformatted long data
• Short - Reformatted short data
• Stocks FNO - FNO stock list
• BUY - Only BUY signals (Green)
• SELL - Only SELL signals (Red)
• FNO - FNO analysis

Best regards,
Chartink Automation
"""
        msg.attach(MIMEText(body, 'plain'))

        # Attach file
        with open(file_path, 'rb') as f:
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file_path)}')
        msg.attach(attachment)

        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()

        print(f"[EMAIL] ✅ Sent successfully to {EMAIL_RECEIVER}")
        return True

    except Exception as e:
        print(f"[ERROR] Email failed: {e}")
        print("[TIP] For Outlook: Use your normal password (NOT App Password)")
        print("[TIP] For Gmail: Use App Password (16-char)")
        return False

def main():
    if SKIP_WEEKEND and not is_weekday():
        print("[INFO] Weekend (Sat/Sun) - Skipping run")
        print("[INFO] To run on weekends, set SKIP_WEEKEND=false in secrets")
        print("=" * 70)
        print("  ⏭️ SKIPPED - Weekend")
        print("=" * 70)
        return

    print("=" * 70)
    print("  CHARTINK DASHBOARD AUTOMATION")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Weekday: {datetime.now().strftime('%A')}")
    print(f"  Email Provider: {EMAIL_PROVIDER}")
    print("=" * 70)

    create_all_folders()

    print("\n[PHASE 1] Downloading from Chartink...")
    download_files()

    print("\n[PHASE 2] Processing...")
    chartink1_df = process_chartink1_folder()
    chartink_df = process_chartink_folder()

    print("\n[PHASE 3] Creating sheets...")
    long_df = create_long_sheet(chartink1_df)
    short_df = create_short_sheet(chartink_df)
    buy_df = create_buy_sheet(long_df)
    sell_df = create_sell_sheet(short_df)
    fno_df = create_fno_sheet(chartink1_df, chartink_df)

    print("\n[PHASE 4] Generating Dashboard...")
    output_file = generate_dashboard(chartink1_df, chartink_df, long_df, short_df, buy_df, sell_df, fno_df)

    print("\n[PHASE 5] Sending Email...")
    send_email(output_file)

    print("\n" + "=" * 70)
    print("  ✅ ALL DONE!")
    print("=" * 70)

if __name__ == "__main__":
    main()
